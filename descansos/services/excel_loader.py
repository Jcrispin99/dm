"""Lee un Excel de descansos médicos y lo persiste en la BD.

Reglas:
- mapeo de columnas por nombre (ver column_mapper)
- los días se recalculan como (fecha_fin - fecha_inicio).days + 1
- dedup por (paciente, fecha_inicio, fecha_fin) → si ya existe, se omite
- si una fila es inválida se reporta pero no rompe la carga
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime

from django.db import transaction
from openpyxl import load_workbook

from ..models import DescansoMedico, Gerencia, Motivo, Paciente
from .column_mapper import find_header_in_sheet


@dataclass
class LoadResult:
    archivo: str = ''
    leidas: int = 0
    insertadas: int = 0
    duplicadas: int = 0
    invalidas: int = 0
    errores_columnas: list[str] = field(default_factory=list)
    errores_filas: list[dict] = field(default_factory=list)

    @property
    def exito(self) -> bool:
        return not self.errores_columnas


def _to_date(value) -> date | None:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y'):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _clean(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def cargar_excel(file_obj, nombre_archivo: str = '') -> LoadResult:
    result = LoadResult(archivo=nombre_archivo)

    wb = load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb.active

    # Recorremos la hoja una sola vez y guardamos las filas; así podemos
    # escanear las primeras buscando la cabecera (puede estar en F5, B3,
    # donde sea) y luego iterar desde la fila inmediatamente debajo.
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        result.errores_columnas.append('El Excel está vacío.')
        return result

    header_idx, mapping, missing, header_row = find_header_in_sheet(iter(all_rows))

    if header_idx == -1 or missing:
        if header_idx == -1:
            result.errores_columnas.append(
                'No se encontró ninguna fila de cabecera con las columnas esperadas '
                '(Código, Nombre del Trabajador, Gerencia, Fecha de Inicio DM, '
                'Fecha de Término DM, Clasificación) en las primeras 50 filas.'
            )
        else:
            detectados = [h for h in header_row if h not in (None, '')]
            result.errores_columnas.append(
                'Faltan columnas obligatorias: ' + ', '.join(missing) +
                f'. Cabecera detectada en fila {header_idx + 1}: ' +
                ', '.join(str(d) for d in detectados)
            )
        return result

    data_rows = all_rows[header_idx + 1:]
    base_row_number = header_idx + 2  # número de fila Excel (1-based) de la primera fila de datos

    with transaction.atomic():
        for offset, row in enumerate(data_rows):
            excel_row_num = base_row_number + offset
            if row is None or all(cell is None or cell == '' for cell in row):
                continue

            result.leidas += 1

            def get(field_name: str):
                idx = mapping.get(field_name)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            codigo = _clean(get('codigo'))
            nombre = _clean(get('nombre'))
            gerencia_nombre = _clean(get('gerencia'))
            fecha_inicio = _to_date(get('fecha_inicio'))
            fecha_fin = _to_date(get('fecha_fin'))
            motivo_nombre = _clean(get('motivo'))
            observaciones = _clean(get('observaciones'))

            problemas = []
            if not codigo:
                problemas.append('código vacío')
            if not nombre:
                problemas.append('nombre vacío')
            if not gerencia_nombre:
                problemas.append('gerencia vacía')
            if not motivo_nombre:
                problemas.append('clasificación vacía')
            if fecha_inicio is None:
                problemas.append('fecha de inicio inválida')
            if fecha_fin is None:
                problemas.append('fecha de término inválida')
            if fecha_inicio and fecha_fin and fecha_fin < fecha_inicio:
                problemas.append('fecha_fin < fecha_inicio')

            if problemas:
                result.invalidas += 1
                result.errores_filas.append({
                    'fila': excel_row_num,
                    'problemas': problemas,
                })
                continue

            paciente, _ = Paciente.objects.get_or_create(
                codigo=codigo, defaults={'nombre': nombre},
            )
            if paciente.nombre != nombre and nombre:
                paciente.nombre = nombre
                paciente.save(update_fields=['nombre'])

            gerencia, _ = Gerencia.objects.get_or_create(nombre=gerencia_nombre)
            motivo, _ = Motivo.objects.get_or_create(nombre=motivo_nombre)

            exists = DescansoMedico.objects.filter(
                paciente=paciente,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
            ).exists()
            if exists:
                result.duplicadas += 1
                continue

            DescansoMedico.objects.create(
                paciente=paciente,
                gerencia=gerencia,
                motivo=motivo,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                observaciones=observaciones,
                archivo_origen=nombre_archivo,
            )
            result.insertadas += 1

    return result
